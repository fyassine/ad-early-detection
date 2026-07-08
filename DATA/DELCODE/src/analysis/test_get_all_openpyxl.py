import openpyxl
from pathlib import Path
import json

REPO_ROOT = Path('/mnt/e/fyassine/ad-early-detection')
METADATA_DIR = REPO_ROOT / "DATA" / "DELCODE" / "__metadata__"
DOCS_DIR = METADATA_DIR / "documentation"

BL_EXCEL = DOCS_DIR / "Antrag_462_Ruat_DysConnectivity Index_20240320_DELCODE_Baseline_repseudonymisiert.xlsx"
FU_EXCEL = DOCS_DIR / "Antrag_462_Ruat_DysConnectivity Index_20240320DELCODE_Follow-up_repseudonymisiert.xlsx"
EXTRA_EXCEL = DOCS_DIR / "Antrag 462_Ruat_MRT-Subjects_nicht-freigegeben.xlsx"

def get_all_documented_pseudonyms() -> set:
    files = [BL_EXCEL, FU_EXCEL, EXTRA_EXCEL]
    all_pseudonyms = set()
    for file in files:
        if not file.exists():
            continue
        print(f"Reading {file.name}")
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            # Find the pseudonym column
            col_indices = []
            header_found = False
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                for idx, cell_value in enumerate(row):
                    if cell_value and str(cell_value).lower().strip() in ('pseudonym', 'repseudonym'):
                        col_indices.append(idx)
                        header_found = True
                if header_found:
                    break
            
            if not col_indices:
                continue
                
            # Now read those columns
            for row in ws.iter_rows(min_row=1, values_only=True):
                for idx in col_indices:
                    if idx < len(row):
                        val = row[idx]
                        if val is not None:
                            val_str = str(val).strip()
                            if val_str and val_str.lower() not in ('nan', 'none', '', 'pseudonym', 'repseudonym'):
                                all_pseudonyms.add(val_str)
        wb.close()
    return all_pseudonyms

print(len(get_all_documented_pseudonyms()))
